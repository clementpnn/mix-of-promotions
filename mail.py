# Import des librairies
# pip install email-to
import email_to
# pip install openpyxl
from openpyxl import load_workbook

# Docs :
# https://email-to.readthedocs.io/en/latest/readme.html
# https://www.youtube.com/watch?v=g_j6ILT-X0k

def script():
  try:
      # Récupère le fichier
      wb_result = load_workbook('Result.xlsx')

      # Récupère la feuille du fichier
      ws_result = wb_result.active
  except Exception:
    return 'Le fichier Result.xlsx est manquant.'

  # Adresse email d'envoi
  server = email_to.EmailServer('smtp.gmail.com', 587, 'EMAIL QUI ENVOI', 'CODE POUR APP (VOIR VIDEO DANS DOC)')

  # Initialise la ligne et la colonne qui vont définir la cellule sur laquelle lire
  row = 2
  column = 1

  # Lis les données tant qu'il y a des groupes
  while column <= ws_result.max_column:
    name_web_1 = ws_result.cell(row=row, column=column).value.split()
    email_web_1 = ws_result.cell(row=row, column=column+1).value
    name_web_2 = ws_result.cell(row=row+1, column=column).value.split()
    email_web_2 = ws_result.cell(row=row+1, column=column+1).value

    # Vérifie si les cellules ne sont pas vide
    if name_web_1 and email_web_1 and name_web_2 and email_web_2:

      # Le mail
      message = server.message()
      message.add(f'Bonjour {name_web_1[1]} {name_web_1[0]},\n')

      message.add('Cordialement,\n')

      # Le style du mail
      message.style = 'p { color: black}'

      # Adresse email qui reçoi
      message.send(f'{email_web_1}', 'OBJET')

      # Print pour vérifier que tout se passe bien
      print('email envoyé à : ', email_web_1)
    
    # Décale la ligne qui définit la cellule sur laquelle lire
    row +=4

    # Décale de colonne si le texte descend trop bas
    if row >= 120:
      row = 2
      column += 4

script()