# Import des librairies
from random import choice
# pip install openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

def script():

  try:
    # Récupère les fichiers
    wb_1 = load_workbook('Classeur1.xlsx')
    wb_2 = load_workbook('Classeur2.xlsx')
    wb_result = load_workbook('Result.xlsx')

    # Récupère les feuilles des fichiers
    ws_1 = wb_1.active
    ws_2 = wb_2.active
    ws_result = wb_result.active
  except Exception:
    return "L'un des fichiers suivant est manquant : Classeur1.xlsx, Classeur2.xlsx ou Result.xlsx."

  # Supprime totalement l'ancienne feuille
  wb_result.remove(ws_result)
  wb_result.create_sheet(title='Feuil1')
  ws_result = wb_result.active

  # Listes des élèves
  list_1 = [[ws_1.cell(row=i, column=1).value, ws_1.cell(row=i, column=2).value, 'web1'] for i in range(1, ws_1.max_row)]
  list_2 = [[ws_2.cell(row=j, column=1).value, ws_2.cell(row=j, column=2).value, 'web2']for j in range(1, ws_2.max_row)]
  list_group = []

  # Ajoute les élèves au hasard dans la liste groupe et les enlève de leur ancienne liste
  def select_name(list_1, list_2):
    name_1 = choice(list_1)
    list_1.remove(name_1)
    name_2 = choice(list_2)
    list_2.remove(name_2)
    list_group.append([name_1, name_2])

  # Exécute la fonction tant qu'il y a suffisamment d'élèves dans les deux listes
  while list_1 and list_2:
    select_name(list_1, list_2)

  # Ajoute les dernières personnes dans la liste groupe
  def select_last_name(list_last, list_group):
    while list_last:
      if len(list_last) == 3:
        list_group.append(list_last)
        break
      else:
        select_name(list_last, list_last)   
  select_last_name(list_1, list_group)
  select_last_name(list_2, list_group)

  # Initialise la ligne et la colonne qui vont définir la cellule sur laquelle écrire
  row = 1
  column = 1

  # Ajoute une largeur à la colonne souhaitée via la liste de lettres
  def width_col(col):
    list_col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'BB', 'CC', 'DD', 'EE', 'FF', 'GG', 'HH', 'II', 'JJ', 'KK', 'LL', 'MM', 'NN', 'OO', 'PP', 'QQ', 'RR', 'SS', 'TT', 'UU', 'VV', 'WW', 'XX', 'YY', 'ZZ']
    ws_result.column_dimensions[list_col[col]].width = 17
  width_col(0)
  width_col(1)

  # On récupère un groupe dans la liste de groupes et son numéro automatiquement incrémenté
  for nb_group, group in enumerate(list_group, start=1):

    # Stylise la ligne qui indique le numéro de groupe
    ws_result.merge_cells(start_row = row, start_column = column, end_row = row, end_column = column+2)
    ws_result.cell(row=row, column=column).value = f'GROUPE {nb_group}'
    ws_result.cell(row=row, column=column).alignment = Alignment(horizontal = 'center')
    ws_result.cell(row=row, column=column).fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    
    # Ajoute les membres du groupe et leurs emails
    ws_result.cell(row=row+1, column=column).value = group[0][0]
    ws_result.cell(row=row+1, column=column+1).value = group[0][1]
    ws_result.cell(row=row+1, column=column+2).value = group[0][2]
    ws_result.cell(row=row+2, column=column).value = group[1][0]
    ws_result.cell(row=row+2, column=column+1).value = group[1][1]
    ws_result.cell(row=row+2, column=column+2).value = group[1][2]

    # Ajoute le dernier membre si le groupe est constitué de 3 personnes
    if len(group) != 2:
      ws_result.cell(row=row+3, column=column).value = group[2][0]
      ws_result.cell(row=row+3, column=column+1).value = group[2][1]
      ws_result.cell(row=row+3, column=column+2).value = group[2][2]
    
    # Décale la ligne qui définit la cellule sur laquelle écrire
    row +=4

    # Décale de colonne si le texte descend trop bas
    if row >= 120:
      row = 1
      column += 4
      width_col(column-1)
      width_col(column)

  try:
    wb_result.save('Result.xlsx')
    wb_result.close()
    print('Le traitement des données est terminé.')
  except Exception:
    return 'Erreur de sauvegarde, fermez le fichier "Result.xlsx" et relancer le script'

script()